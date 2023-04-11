# Importing the modules
import csv
import openpyxl
import os
import tkinter as tk
from tkinter import messagebox
from tkinter import * 
from tkinter.ttk import *
from tkinter import filedialog


def cleanArray(array_di_stringhe):
    for i in range(len(array_di_stringhe)):
        if "[" in str(array_di_stringhe[i]):
            array_di_stringhe[i] = array_di_stringhe[i].replace("[", "\[")
        if "]" in str(array_di_stringhe[i]):
            array_di_stringhe[i] = array_di_stringhe[i].replace("]", "\]")
    return array_di_stringhe


def datiAnagrafica(fileIn, fileOut):
    # try:
        # Definisci gli header per il tuo file CSV
    headers = ['ID', 'Reperto', 'Materiale', 'Altezza', 'Lunghezza', 'Larghezza', 'Spessore', 'Diametro', 'Stato di Conservazione', 'Descrizione', 'Note',
            'Soggetto Iconografico', 'Cronologia', 'Nazione', 'Città', 'Status', 'Museo / Collezionista', '# inv.', 'Data di Acquisizione', 'Modalità di Acquisizione']

    # Apri il file CSV in modalità scrittura
    directory = "data"  # Specifico la directory
    filename = fileOut  # Specifico il nome del file

    if not os.path.exists(directory):  # Creo la directory se non esiste
        os.makedirs(directory)

    filePathAndName = os.path.join(
        directory, filename)  # Creo il full name path

    with open(filePathAndName, 'w', newline='', encoding="utf-8") as file:
        # Creo un oggetto writer per scrivere nel file CSV
        writer = csv.writer(file)
        # Apro il file Excel
        workbook = openpyxl.load_workbook(fileIn)

        # Seleziono il foglio di lavoro che vuoi leggere
        worksheet = workbook['Dati Anagrafici']

        # Scrivo gli header nel file CSV
        writer.writerow(headers)

        # Leggo i dati nell'excel ciclicamente e li salvo in un oggetto per poi metterli nel CSV
        for row in worksheet.iter_rows(min_row=3, min_col=1, max_col=21):

            # Creo una lista vuota per contenere il contenuto delle celle
            cell_content = []

            # Itero su ogni cella e aggiungi il contenuto alla lista
            for cell in row:
                cell_content.append(cell.value)

            # Stampo il contenuto
            # print (cell_content)

            # Se c'è qualche valore nella riga allora:
            if (cell_content[0]):
                # Sostituisco il contenuto Null con la stringa 'Missing Value'
                for i in range(len(cell_content)):
                    if cell_content[i] is None:
                        cell_content[i] = 'Missing Value'

                # Rimuovo eventuali caratteri che possono dare errori
                # cell_content = cleanArray(cell_content)

                # Scrivo il contenuto
                writer.writerow([cell_content[0], cell_content[2], cell_content[3], cell_content[4], cell_content[5], cell_content[6], cell_content[7], cell_content[8], cell_content[9], cell_content[10],
                                cell_content[11], cell_content[12], cell_content[13], cell_content[14], cell_content[15], cell_content[16], cell_content[17], cell_content[18], cell_content[19], cell_content[20]])
    # except:
    #     messagebox.showinfo('Errore','C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')
    #     print('C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')




def datiScavo(fileIn, fileOut):
    try:
        # Definisci gli header per il tuo file CSV
        headers = ['ID','ID SCAVO','Toponimo','Regio','Insula','Civico','Stanza','Informazioni di Dettaglio','Indicazioni Generali','Giorno','Mese','Anno','Soprintendente','Architetto Direttore','Note','Archivio','Segnatura','Riferimento PAH','Citazione']

        # Apri il file CSV in modalità scrittura
        directory = "data"  # Specifico la directory
        filename = fileOut  # Specifico il nome del file

        if not os.path.exists(directory):  # Creo la directory se non esiste
            os.makedirs(directory)

        filePathAndName = os.path.join(
            directory, filename)  # Creo il full name path

        with open(filePathAndName, 'w', newline='', encoding="utf-8") as file:
            # Creo un oggetto writer per scrivere nel file CSV
            writer = csv.writer(file)
            # Apro il file Excel
            workbook = openpyxl.load_workbook(fileIn)

            # Seleziono il foglio di lavoro che vuoi leggere
            worksheet = workbook['Dati Scavo']

            # Scrivo gli header nel file CSV
            writer.writerow(headers)

            # Leggo i dati nell'excel ciclicamente e li salvo in un oggetto per poi metterli nel CSV
            for row in worksheet.iter_rows(min_row=3, min_col=1, max_col=19):

                # Creo una lista vuota per contenere il contenuto delle celle
                cell_content = []

                # Itero su ogni cella e aggiungi il contenuto alla lista
                for cell in row:
                    cell_content.append(cell.value)

                # Stampo il contenuto
                # print (cell_content)

                # Se c'è qualche valore nella riga allora:
                if (cell_content[0]):
                    # Sostituisco il contenuto Null con la stringa 'Missing Value'
                    for i in range(len(cell_content)):
                        if cell_content[i] is None:
                            cell_content[i] = 'Missing Value'
                    # Scrivo il contenuto
                    writer.writerow([cell_content[0], float(cell_content[1]), cell_content[2], cell_content[3], cell_content[4], cell_content[5], cell_content[6], cell_content[7], cell_content[8], cell_content[9], cell_content[10],
                                cell_content[11], cell_content[12], cell_content[13], cell_content[14], cell_content[15], cell_content[16], cell_content[17], cell_content[18]])
    except:
        messagebox.showinfo('Errore','C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')
        print('C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')


def datiCollezionistici(fileIn, fileOut):
    try:
        # Definisci gli header per il tuo file CSV
        headers = ['ID','ID COLLEZIONISTA','Nazione','Città','Collezionista','Luogo','# inv.','da','a','Modalità di Acquisizione','Nazione 2','Città 2','Venditore','Nome','Note','Archivio','Segnatura','Documento']

        # Apri il file CSV in modalità scrittura
        directory = "data"  # Specifico la directory
        filename = fileOut  # Specifico il nome del file

        if not os.path.exists(directory):  # Creo la directory se non esiste
            os.makedirs(directory)

        filePathAndName = os.path.join(
            directory, filename)  # Creo il full name path

        with open(filePathAndName, 'w', newline='', encoding="utf-8") as file:
            # Creo un oggetto writer per scrivere nel file CSV
            writer = csv.writer(file)
            # Apro il file Excel
            workbook = openpyxl.load_workbook(fileIn)

            # Seleziono il foglio di lavoro che vuoi leggere
            worksheet = workbook['Dati Collezionistici']

            # Scrivo gli header nel file CSV
            writer.writerow(headers)

            # Leggo i dati nell'excel ciclicamente e li salvo in un oggetto per poi metterli nel CSV
            for row in worksheet.iter_rows(min_row=3, min_col=1, max_col=19):

                # Creo una lista vuota per contenere il contenuto delle celle
                cell_content = []

                # Itero su ogni cella e aggiungi il contenuto alla lista
                for cell in row:
                    cell_content.append(cell.value)

                # Stampo il contenuto
                # print (cell_content)

                # Se c'è qualche valore nella riga allora:
                if (cell_content[0]):
                    # Sostituisco il contenuto Null con la stringa 'Missing Value'
                    for i in range(len(cell_content)):
                        if cell_content[i] is None:
                            cell_content[i] = 'Missing Value'
                    # Rimuovo eventuali caratteri che possono dare errori
                    # cell_content = cleanArray(cell_content)
                    # Scrivo il contenuto
                    writer.writerow([cell_content[0], float(cell_content[1]), cell_content[2], cell_content[3], cell_content[4], cell_content[5], cell_content[6], cell_content[7], cell_content[8], cell_content[9], cell_content[10],
                                cell_content[11], cell_content[12], cell_content[13], cell_content[14], cell_content[15], cell_content[16], cell_content[17]])
    except:
        messagebox.showinfo('Errore','C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')
        print('C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')





def datiBibliografici(fileIn, fileOut):
    try:
        # Definisci gli header per il tuo file CSV
        headers = ['ID','Bibliografia']

        # Apri il file CSV in modalità scrittura
        directory = "data"  # Specifico la directory
        filename = fileOut  # Specifico il nome del file

        if not os.path.exists(directory):  # Creo la directory se non esiste
            os.makedirs(directory)

        filePathAndName = os.path.join(
            directory, filename)  # Creo il full name path

        with open(filePathAndName, 'w', newline='', encoding="utf-8") as file:
            # Creo un oggetto writer per scrivere nel file CSV
            writer = csv.writer(file)
            # Apro il file Excel
            workbook = openpyxl.load_workbook(fileIn)

            # Seleziono il foglio di lavoro che vuoi leggere
            worksheet = workbook['Dati Bibliografici']

            # Scrivo gli header nel file CSV
            writer.writerow(headers)

            # Leggo i dati nell'excel ciclicamente e li salvo in un oggetto per poi metterli nel CSV
            for row in worksheet.iter_rows(min_row=3, min_col=1, max_col=2):

                # Creo una lista vuota per contenere il contenuto delle celle
                cell_content = []

                # Itero su ogni cella e aggiungi il contenuto alla lista
                for cell in row:
                    cell_content.append(cell.value)

                # Stampo il contenuto
                # print (cell_content)

                # Se c'è qualche valore nella riga allora:
                if (cell_content[0]):
                    # Sostituisco il contenuto Null con la stringa 'Missing Value'
                    for i in range(len(cell_content)):
                        if cell_content[i] is None:
                            cell_content[i] = 'Missing Value'

                    # Rimuovo eventuali caratteri che possono dare errori
                    cell_content = cleanArray(cell_content)
                    # Scrivo il contenuto
                    writer.writerow([cell_content[0], cell_content[1]])
    except:
        messagebox.showinfo('Errore','C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')
        print('C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')


def abbreviazioniArchivi(fileIn, fileOut):
    try:
        # Definisci gli header per il tuo file CSV
        headers = ['','Abbreviazione','Archivio','']

        # Apri il file CSV in modalità scrittura
        directory = "data"  # Specifico la directory
        filename = fileOut  # Specifico il nome del file

        if not os.path.exists(directory):  # Creo la directory se non esiste
            os.makedirs(directory)

        filePathAndName = os.path.join(
            directory, filename)  # Creo il full name path

        with open(filePathAndName, 'w', newline='', encoding="utf-8") as file:
            # Creo un oggetto writer per scrivere nel file CSV
            writer = csv.writer(file)
            # Apro il file Excel
            workbook = openpyxl.load_workbook(fileIn)

            # Seleziono il foglio di lavoro che vuoi leggere
            worksheet = workbook['Abbreviazioni Archivi']

            # Scrivo gli header nel file CSV
            writer.writerow(headers)

            #Inizializzo un contatore che stamperò nella prima colonna e mi serve da indice
            contatore = 0

            # Leggo i dati nell'excel ciclicamente e li salvo in un oggetto per poi metterli nel CSV
            for row in worksheet.iter_rows(min_row=0, min_col=0, max_col=3):

                # Creo una lista vuota per contenere il contenuto delle celle
                cell_content = []

                # Itero su ogni cella e aggiungi il contenuto alla lista
                for cell in row:
                    cell_content.append(cell.value)

                # Stampo il contenuto
                # print (cell_content)

                # Se c'è qualche valore nella riga allora:
                if (cell_content[0]):
                    # Sostituisco il contenuto Null con la stringa 'Missing Value'
                    for i in range(len(cell_content)):
                        if cell_content[i] is None:
                            cell_content[i] = 'Missing Value'

                    # Rimuovo eventuali caratteri che possono dare errori
                    # cell_content = cleanArray(cell_content)
                    # Scrivo il contenuto
                    writer.writerow([contatore, cell_content[0], cell_content[1], ''])
                    contatore += 1
    except:
        messagebox.showinfo('Errore','C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')
        print('C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')


def abbreviazioniBibliografiche(fileIn, fileOut):
    try:
        # Definisci gli header per il tuo file CSV
        headers = ['','Abbreviazione','Bibliografia','']

        # Apri il file CSV in modalità scrittura
        directory = "data"  # Specifico la directory
        filename = fileOut  # Specifico il nome del file

        if not os.path.exists(directory):  # Creo la directory se non esiste
            os.makedirs(directory)

        filePathAndName = os.path.join(
            directory, filename)  # Creo il full name path

        with open(filePathAndName, 'w', newline='', encoding="utf-8") as file:
            # Creo un oggetto writer per scrivere nel file CSV
            writer = csv.writer(file)
            # Apro il file Excel
            workbook = openpyxl.load_workbook(fileIn)

            # Seleziono il foglio di lavoro che vuoi leggere
            worksheet = workbook['Abbreviazioni Bibliografiche']

            # Scrivo gli header nel file CSV
            writer.writerow(headers)

            #Inizializzo un contatore che stamperò nella prima colonna e mi serve da indice
            contatore = 0

            # Leggo i dati nell'excel ciclicamente e li salvo in un oggetto per poi metterli nel CSV
            for row in worksheet.iter_rows(min_row=0, min_col=0, max_col=3):

                # Creo una lista vuota per contenere il contenuto delle celle
                cell_content = []

                # Itero su ogni cella e aggiungi il contenuto alla lista
                for cell in row:
                    cell_content.append(cell.value)

                # Stampo il contenuto
                # print (cell_content)

                # Se c'è qualche valore nella riga allora:
                if (cell_content[0]):
                    # Sostituisco il contenuto Null con la stringa 'Missing Value'
                    for i in range(len(cell_content)):
                        if cell_content[i] is None:
                            cell_content[i] = 'Missing Value'

                    # Rimuovo eventuali caratteri che possono dare errori
                    # cell_content = cleanArray(cell_content)
                    # Scrivo il contenuto
                    writer.writerow([contatore, cell_content[0], cell_content[1], ''])
                    contatore += 1
    except:
        messagebox.showinfo('Errore','C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')
        print('C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')


def abbreviazioniTipologiche(fileIn, fileOut):
    try:
        # Definisci gli header per il tuo file CSV
        headers = ['','Abbreviazione','Tipo','']

        # Apri il file CSV in modalità scrittura
        directory = "data"  # Specifico la directory
        filename = fileOut  # Specifico il nome del file

        if not os.path.exists(directory):  # Creo la directory se non esiste
            os.makedirs(directory)

        filePathAndName = os.path.join(
            directory, filename)  # Creo il full name path

        with open(filePathAndName, 'w', newline='', encoding="utf-8") as file:
            # Creo un oggetto writer per scrivere nel file CSV
            writer = csv.writer(file)
            # Apro il file Excel
            workbook = openpyxl.load_workbook(fileIn)

            # Seleziono il foglio di lavoro che vuoi leggere
            worksheet = workbook['Abbreviazioni Tipologiche']

            # Scrivo gli header nel file CSV
            writer.writerow(headers)

            #Inizializzo un contatore che stamperò nella prima colonna e mi serve da indice
            contatore = 0

            # Leggo i dati nell'excel ciclicamente e li salvo in un oggetto per poi metterli nel CSV
            for row in worksheet.iter_rows(min_row=0, min_col=0, max_col=3):

                # Creo una lista vuota per contenere il contenuto delle celle
                cell_content = []

                # Itero su ogni cella e aggiungi il contenuto alla lista
                for cell in row:
                    cell_content.append(cell.value)
                    
                # Se c'è qualche valore nella riga allora:
                if (cell_content[0]):
                    # Sostituisco il contenuto Null con la stringa 'Missing Value'
                    for i in range(len(cell_content)):
                        if cell_content[i] is None:
                            cell_content[i] = 'Missing Value'

                    # Rimuovo eventuali caratteri che possono dare errori
                    # cell_content = cleanArray(cell_content)
                    # Scrivo il contenuto
                    writer.writerow([contatore, cell_content[0], cell_content[1], ''])
                    contatore += 1
    except: 
        messagebox.showinfo('Errore','C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')
        print('C\'è stato un errore duranta la creazione del file: \"' + fileOut + '\"')


def execute(excelFile):
    # Progress bar widget
    progress = Progressbar(window, orient = HORIZONTAL, length = 100, mode = 'determinate')
    progress.pack(pady = 0)
    try:
        progress['value'] = 5
        window.update_idletasks()
        datiAnagrafica(excelFile, 'dati_anagrafica_clean.csv')
        progress['value'] = 10
        window.update_idletasks()
        datiScavo(excelFile, 'dati_scavo_clean.csv')
        progress['value'] = 20
        window.update_idletasks()
        datiCollezionistici(excelFile, 'dati_collezionisti_clean.csv')
        progress['value'] = 30
        window.update_idletasks()
        datiBibliografici(excelFile, 'dati_bibliografici_clean.csv')
        progress['value'] = 40
        window.update_idletasks()
        abbreviazioniArchivi(excelFile, 'abbreviazioni_archivi.csv')
        progress['value'] = 80
        window.update_idletasks()
        abbreviazioniBibliografiche(excelFile, 'abbreviazioni_bibliografia.csv')
        progress['value'] = 90
        window.update_idletasks()
        abbreviazioniTipologiche(excelFile, 'abbreviazioni_tipologie.csv')
        progress['value'] = 100
        window.update_idletasks()
        progress.destroy()
        window.update_idletasks()
        label = tk.Label(window, text="Esportazione completata", fg='green', font=('Times', 24))
        label.pack()
        sottotitolo = tk.Label(window, text="Riavvia il programma per una nuova operazione", fg='grey', font=('Times', 8))
        sottotitolo.pack()
    except: 
        quit()




# Creo una nuova finestra
window = tk.Tk()

# Titolo della finestra
window.title("Memorabilia Pompeiana")

# Dimensioni della finestra
window.geometry("800x200")

# Etichetta alla finestra vuota per fare spazio (sono un programmatore scarso)
label = tk.Label(window, text=" ")
label.pack()

# Import required libraries
from tkinter import *
from tkinter import filedialog

# Create a dialog using filedialog function
window.filename=filedialog.askopenfilename(initialdir=os.getcwd(), title="Select a file", filetypes=[("Excel files", ".xlsx"),("all files", "*.*")])

# Create a label widget
try:
    label=Label(window, text="Il file selezionato è: \"" + window.filename + "\"", font=('Times', 12))
    label.pack()
except:
    quit()


# Etichetta alla finestra vuota per fare spazio (sono un programmatore scarso x2)
label = tk.Label(window, text=" ")
label.pack()

# Pulsante alla finestra
button = tk.Button(window, text="Clicca qui", command=execute(window.filename))
button.pack()
button.destroy()

# Etichetta alla finestra vuota per fare spazio (sono un programmatore scarso x3)
label = tk.Label(window, text=" ")
label.pack()

# Avvio la finestra
window.mainloop()